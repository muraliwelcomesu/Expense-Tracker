#! python3
from tkinter import * 
from tkinter import ttk
from tkinter import messagebox
import openpyxl,os, sys,webbrowser,calendar
import pandas as pd
import Utils_Common as utils
import Launch_Browser
from datetime import datetime
from tkinter import Menu
import User 

def _quit():
    root.quit()
    root.destroy()
    exit()

def fn_prepare_sheet(Path,FileName,SheetName):
    try:
        #print('inside fn_prepare_sheet')
        l_Ref_Sheet_Name = User.Exp_MasterRec_sheetName
        RefExcel =  User.Password_Excel 
        wb = openpyxl.load_workbook(RefExcel)
        Ref_sheet = wb[l_Ref_Sheet_Name]
    
        l_file_found = utils.CheckFile(Path,FileName)
        #print('l_file_found'+str(l_file_found))
        #print(l_file_found)
        if l_file_found < 1:
            #print('Will create new file')
            wb1 = openpyxl.Workbook()
            wrk_sheet = wb1.create_sheet()
            wrk_sheet.title = (SheetName.upper()).strip()
            wrk_sheet = utils.PrepareSheet(Ref_sheet,wrk_sheet)
            wb1.save(FileName)
        else:
            #print('file available..check if sheet is available..')
            wrk_sheet = utils.getExcelSheet(Path,FileName,SheetName)
            if wrk_sheet is None:
                #print('Creating sheet')
                wb1 = openpyxl.load_workbook(FileName)
                wrk_sheet = wb1.create_sheet()
                wrk_sheet.title =(SheetName.upper()).strip()
                wrk_sheet = utils.PrepareSheet(Ref_sheet,wrk_sheet)
                wb1.save(FileName)
            #else:
                #print('Sheet already available')
    except:
        messagebox.showinfo(title='Error Message', message = 'Error in Preparing/Saving excel')
          
        

def fn_launch_url(fn_name):
    print(fn_name)
    if ',' in fn_name:
        l_list = fn_name.split(',')
        for i in l_list:
            l=messagebox.askyesno("Confirm",'Launch  :'+i)
            if l:
                result = getattr(Launch_Browser.OpenURL, i)()
    else:
        if fn_name != 'N':
            l=messagebox.askyesno("Confirm",'Launch :'+fn_name)
            if l:
                result = getattr(Launch_Browser.OpenURL, fn_name)()

def get_data_excel():
    #print('insdie get data excel')
    p_month = month.get().upper()
    p_year = year.get()
    Path = User.Expense_Track_Path
    FileName = User.Expense_Track_ExcelName
    SheetName = p_month+p_year;
    fn_prepare_sheet(Path,FileName,SheetName)
    #print('after prepare sheet')
    os.chdir(Path)
    df_tmp = pd.read_excel(FileName,sheet_name = SheetName) 
    df_tmp.set_index("Name",inplace = True)
 
    return df_tmp
    
def fn_upd_status(row,index):
    #print('inside upd_Status')
    #print(row["Name"])
    Path = User.Expense_Track_Path
    p_month = month.get().upper()
    p_year = year.get()
    FileName = User.Expense_Track_ExcelName
    SheetName = p_month+p_year;
    df = get_data_excel()
    df.loc[index,'Status'] = 'Paid'
    utils.fn_save_Excel(FileName,df,SheetName)

        
def fn_btn_Modify():
    #print('Modify Entry')
    Path = User.Expense_Track_Path
    FileName = User.Expense_Track_ExcelName
    os.chdir(Path)


    def fn_fetch_Rec(df1,p_list):
        #print('inside fn_fetch_rec')
        lbl_Duedate = ttk.Label(frame_body,text = "DueDate :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 2,column =0)
        txt_Duedate =  ttk.Entry(frame_body,width = 20)
        txt_Duedate.insert(0, (df1.loc[p_list].DueDate))
        txt_Duedate.grid(row = 2,column = 1)
        lbl_Amount = ttk.Label(frame_body,text = "Amount  :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 3,column = 0)
        txt_Amount =  ttk.Entry(frame_body,width = 20)
        txt_Amount.insert(0, (df1.loc[p_list].Amount))
        txt_Amount.grid(row = 3,column = 1)
        
        lbl_Status = ttk.Label(frame_body,text = "Status  :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 4,column = 0)
        txt_Status =  ttk.Entry(frame_body,width = 20)
        selopt = df1.loc[p_list].Status
        txt_Status.insert(0,selopt)
        txt_Status.grid(row = 4,column = 1)

        lbl_Function  = ttk.Label(frame_body,text = "Function",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 5,column = 0)
        txt_Function =  ttk.Entry(frame_body,width = 20)   
        txt_Function.insert(0, df1.loc[p_list]["Function Name"])
        txt_Function.grid(row = 5,column = 1)
            
        def fn_update_entry(df1,p_list):
            df1.at[p_list,"DueDate"]= txt_Duedate.get()
            df1.at[p_list,"Amount"]= txt_Amount.get()
            df1.at[p_list,"Status"]= txt_Status.get()
            df1.at[p_list,"Function Name"]= txt_Function.get()
            SheetName = sheetNames.get()
            utils.fn_save_Excel(FileName,df1,SheetName)
        
        btn_SaveEntry = ttk.Button(frame_body,text = 'Update Entry',width = 20, command = lambda: fn_update_entry(df1,p_list))
        btn_SaveEntry.grid(row = 6 ,column = 0)
        
        

    for widget in frame_body.winfo_children():
        widget.destroy()
    
    p_month = month.get().upper()
    p_year = year.get()
    SheetName = p_month+p_year;
    
    sheetNames = StringVar()
    combobox = ttk.Combobox(frame_body, textvariable = sheetNames)
    l_sheets = []
    l_excel = pd.ExcelFile(FileName)
    for sheet in l_excel.sheet_names:
        l_sheets.append(sheet)
    combobox.config(values = list(l_sheets))
    combobox.grid(row = 0 , column = 1)
        
    lbl_Sheet =  ttk.Label(frame_body,text = "MonthYear[MONYYYY]:",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 0,column = 0)
    lbl_Name =  ttk.Label(frame_body,text = "Name :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 1,column = 0)
    
    Lst_Names = StringVar()
    combobox1 = ttk.Combobox(frame_body, textvariable = Lst_Names)
    l_Names = []
    combobox1.grid(row = 1 , column = 1)
    sheetNames.set(SheetName)
    df1 = pd.read_excel(FileName,sheet_name = sheetNames.get())
    df1.set_index("Name",inplace = True)
    l_Names = df1.index.tolist()
    combobox1.config(values = list(l_Names))
    btn_Fetch = ttk.Button(frame_body,text = 'Fetch',width = 20, command = lambda: fn_fetch_Rec(df1,Lst_Names.get()))
    btn_Fetch.grid(row = 1 ,column = 2)
 

            
    
def fn_btn_Add():
    print('inside fn_btn_Add')
    Path = User.Expense_Track_Path
    FileName = User.Expense_Track_ExcelName
    os.chdir(Path)
    p_month = month.get().upper()
    p_year = year.get()
    SheetName = p_month+p_year;
    for widget in frame_body.winfo_children():
        widget.destroy()
        
    def fn_save_entry():
        SheetName =sheetNames.get()  
        dict1 = {'No':['1'],
                 'Name' :[txt_Name.get()],
                 'DueDate':[txt_Duedate.get()],
                 'Amount':[txt_Amount.get()],
                 'Status':['Pending'],
                 'Holidays':['0'],
                 'Function Name':['N']}
        df_new = pd.DataFrame(data = dict1)
        df_new.set_index("Name",inplace = True)
        #print(df_new)
        #print('printing new dataframe done')
        df1 =    pd.read_excel(FileName,sheet_name = SheetName)
        df1.set_index("Name",inplace = True)
        df2 = df1.append(df_new)#,ignore_index = True)
        utils.fn_save_Excel(FileName,df2,SheetName)
        fn_clear_entry()
       
    def fn_clear_entry():
        txt_Name.delete(0, END)
        txt_Duedate.delete(0, END)
        txt_Amount.delete(0, END)
        txt_Function.delete(0, END)
        #txt_Sheet.delete(0,END)
        
    sheetNames = StringVar()
    combobox = ttk.Combobox(frame_body, textvariable = sheetNames)
    l_sheets = []
    l_excel = pd.ExcelFile(FileName)
    for sheet in l_excel.sheet_names:
        l_sheets.append(sheet)
    combobox.config(values = list(l_sheets))
    combobox.grid(row = 0 , column = 1)
    sheetNames.set(SheetName)     
    lbl_Sheet =  ttk.Label(frame_body,text = "MonthYear[MONYYYY]:",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 0,column = 0)
 
    lbl_Name =  ttk.Label(frame_body,text = "Name :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 1,column = 0)
    txt_Name =  ttk.Entry(frame_body,width = 20)
    txt_Name.grid(row =1,column = 1)
    lbl_Duedate = ttk.Label(frame_body,text = "DueDate :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 2,column =0)
    txt_Duedate =  ttk.Entry(frame_body,width = 20)
    txt_Duedate.grid(row = 2,column = 1)
    lbl_Amount = ttk.Label(frame_body,text = "Amount  :",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 3,column = 0)
    txt_Amount =  ttk.Entry(frame_body,width = 20)
    txt_Amount.grid(row = 3,column = 1)
    lbl_Function  = ttk.Label(frame_body,text = "Function",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = 4,column = 0)
    txt_Function =  ttk.Entry(frame_body,width = 20)    
    txt_Function.grid(row = 4,column = 1)
    btn_SaveEntry = ttk.Button(frame_body,text = 'Save Entry',width = 20, command = lambda: fn_save_entry())
    btn_SaveEntry.grid(row = 5 ,column = 0)
    btn_ClearEntry = ttk.Button(frame_body,text = 'Clear',width = 20, command = lambda: fn_clear_entry())
    btn_ClearEntry.grid(row = 5 ,column = 1)
    
       
def draw_row(frame1,index,no,row,head = 0):

    if head > 0:
        lbl_No =    ttk.Label(frame1,text = 'No',borderwidth=2, relief="solid",width = 5, font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column = 0)
        lbl_Name =  ttk.Label(frame1,text = "Name",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column = 1)
        lbl_Duedate = ttk.Label(frame1,text = "DueDate",borderwidth=2, relief="solid",width = 15,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column =2)
        lbl_Amount = ttk.Label(frame1,text = "Amount",borderwidth=2, relief="solid",width = 15,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column = 3)
        lbl_Status  = ttk.Label(frame1,text = "Status",borderwidth=2, relief="solid",width = 20,font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column = 4)
        lbl_Get = ttk.Label(frame1,text = 'Make Payment',borderwidth=2, relief="solid",width = 20, font = ('Courier',10,'bold'),padding = (3,3)).grid(row = no,column = 5)
        
    else:
        lbl_No =    ttk.Label(frame1,text = no,borderwidth=2, relief="groove",width = 5).grid(row = no, column = 0)
        lbl_Name =  ttk.Label(frame1,text = index,borderwidth=2, relief="groove",width = 20).grid(row = no,column = 1)
        lbl_Duedate = ttk.Label(frame1,text = row["DueDate"],borderwidth=2, relief="groove",width = 15).grid(row = no,column =2)
        lbl_Amount = ttk.Label(frame1,text = row["Amount"],borderwidth=2, relief="groove",width = 15).grid(row = no,column = 3)
        lbl_Status  = ttk.Label(frame1,text = row["Status"],borderwidth=2, relief="groove",width = 20).grid(row = no,column = 4)
        frame_btn = ttk.Frame(frame1)
        btn_View   = ttk.Button(frame_btn,text = 'Fetch',width = 8, command = lambda: fn_launch_url(row["Function Name"]))
        btn_update = ttk.Button(frame_btn,text = 'Update',width = 8, command = lambda: fn_upd_status(row,index))
        if row["Function Name"] == 'N':
            btn_View.state(['disabled'])
        if row["Status"] == 'Paid':
            btn_update.state(['disabled'])
        btn_View.grid(row = 0,column = 1)   
        btn_update.grid(row = 0,column = 2)
        frame_btn.grid(row = no,column = 5) 
  
        btn_Get.grid(row = no,column = 5)   

def fn_btn_Get():
    
    for widget in frame_body.winfo_children():
        widget.destroy()
    l_SelOpt = varSelOpt.get()
    df =  get_data_excel()
         
    if l_SelOpt == 'All':
        l_no = 0
        draw_row(frame_body,'',l_no,'',1)
        for index,row in df.iterrows():
            l_no = l_no + 1
            draw_row(frame_body,index,l_no,row)
    elif l_SelOpt == 'Pending':
        df1 = df[df["Status"] == 'Pending']
        l_no = 0
        draw_row(frame_body,'',l_no,'',1)
        for index,row in df1.iterrows():
            l_no = l_no + 1
            draw_row(frame_body,index,l_no,row)
    elif l_SelOpt == 'Paid':
        df1 = df[df["Status"] == 'Paid']
        l_no = 0
        draw_row(frame_body,'',l_no,'',1)
        for index,row in df1.iterrows():
            l_no = l_no + 1
            draw_row(frame_body,index,l_no,row)     
 

        
    
root = Tk()
root.title('Expense Tracker')

#creating Menu
menubar = Menu()
root.config(menu = menubar)
#add menu items
filemenu = Menu(menubar,tearoff = 0)
filemenu.add_command(label="New")
filemenu.add_separator()
filemenu.add_command(label = "Exit",command = _quit)
menubar.add_cascade(label="File",menu = filemenu)

#add another menu to menubar 

helpmenu = Menu(menubar,tearoff = 0)
helpmenu.add_command(label="About")
helpmenu.add_separator()
menubar.add_cascade(label="Help",menu = helpmenu)


tab_cntl = ttk.Notebook(root)
frame1 = ttk.Frame(tab_cntl)
frame1.config(relief = RIDGE)
frame1.config(padding = (30,15))
tab_cntl.add(frame1,text = 'Expense Tracker')

frame = ttk.Frame(frame1)

frame.config(relief = RIDGE)
frame.config(padding = (30,15))
frame.pack()

frame_body = ttk.Frame(frame1)
frame_body.config(relief = RIDGE)
frame_body.config(padding = (30,15))
frame_body.pack()

frame2 = ttk.Frame(tab_cntl)
frame2.config(relief = RIDGE)
frame2.config(padding = (30,15))
tab_cntl.add(frame2,text = 'Frame2')

tab_cntl.pack(expand = 1,fill = "both")

lbl_Month = ttk.Label(frame,text = 'Month').grid(row = 0 , column = 0)

month = StringVar()
combobox = ttk.Combobox(frame, textvariable = month)
#combobox.pack()
combobox.config(values = ('Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'))
combobox.grid(row = 0 , column = 1)
month.set(datetime.now().strftime("%b")) 
lbl_Year = ttk.Label(frame,text = 'Year').grid(row = 0 , column = 2)
year = StringVar()
Spinbox(frame,from_ = 2017, to = 2025,textvariable = year).grid(row = 0 , column = 3)
l_month = month.get().upper()
year.set('2018')
l_year = year.get()
lbl_Selopt = ttk.Label(frame,text = 'Select').grid(row =1  , column =0)
varSelOpt = StringVar()
comboSel = ttk.Combobox(frame, textvariable = varSelOpt)
#comboSel.pack()
comboSel.config(values = ('All','Pending','Paid'))
#varSelOpt.trace('w', change_varSelOpt)
comboSel.grid(row = 1 , column = 1)
varSelOpt.set('All')

btn_Add = ttk.Button(frame,text='Add Entry' ,command = lambda: fn_btn_Add())
#btn_Add.state(['disabled']) 
btn_Add.grid(row = 2 , column = 1)

btn_Modify = ttk.Button(frame,text='Modify Entry' ,command = lambda: fn_btn_Modify())
#btn_Modify.state(['disabled']) 
btn_Modify.grid(row = 2 , column = 2)

btn_Get = ttk.Button(frame,text='Get Details' ,command = lambda: fn_btn_Get())
#btn_Get.state(['disabled']) 
btn_Get.grid(row = 2 , column = 3)


root.mainloop()
     