#! python3
import openpyxl,os, sys,pyperclip,webbrowser,datetime,calendar
from openpyxl.styles import Font
import pandas as pd
from tkinter import messagebox
#Check the availability of a file in a given location
def CheckFile(path,File):
    os.chdir(path)
    for p_filename in os.listdir():
        if File.upper() == p_filename.upper():
            #print('file found')
            return(int(1))
    #print('file not found')
    return(int(0))

def getExcelSheet(path,fileName,SheetName):
    os.chdir(path)
    l_file_found = CheckFile(path,fileName)
    if l_file_found > 0:
        wb = openpyxl.load_workbook(fileName)
        sheet_list = wb.sheetnames
        if SheetName in sheet_list:
            sheet = wb[SheetName]
            return sheet
        
def fn_save_Excel(FileName,df,SheetName):

    writer = pd.ExcelWriter(FileName)
    l_excel = pd.ExcelFile(FileName)
    for sheet in l_excel.sheet_names:
        if sheet != SheetName:
            df2 = pd.read_excel(FileName,sheet_name = sheet)
            df2.to_excel(writer,sheet)
        else:

            df.to_excel(writer,SheetName)
    try:
        writer.save()
        messagebox.showinfo(title='Success!!', message = 'Successfully Saved')

    except:
        messagebox.showinfo(title='Error Message', message = 'Please close the Excel Files.')
#Prepare Sheet.
def PrepareSheet(ref_sheet,new_sheet):
    #print('Inside PrepareSheet')
    l_new_rownum = 1
    #boldFont =  Font(bold = True)
    l_month_year = (new_sheet.title).strip()
    l_month = l_month_year[:3].capitalize() #Dec
    l_year = l_month_year[3:] #year

    l_months_list = dict((v,k) for k,v in enumerate(calendar.month_abbr))
    new_sheet['A1'] = 'No'
    #new_sheet['A1'].font = boldFont
    new_sheet['B1'] = 'Name'
    new_sheet['C1'] = 'DueDate'
    new_sheet['D1'] = 'Amount'
    new_sheet['E1'] = 'Status'
    new_sheet['F1'] = 'Holidays'
    new_sheet['G1'] = 'Function Name'
    for l_row_num in range(2,ref_sheet.max_row + 1):
        if (str(l_months_list[l_month]).zfill(2) in (ref_sheet['C'+str(l_row_num)].value)  ) or ((ref_sheet['C'+str(l_row_num)].value) == 'ALL'):
            l_new_rownum = l_new_rownum + 1
            for j in range(1,3):
                new_sheet.cell(row=l_new_rownum, column=j).value  = ref_sheet.cell(row=l_row_num, column=j).value
            l_tmp_value  =  str(ref_sheet.cell(row=l_row_num, column=4).value) + '-' + l_month + '-' + l_year
            new_sheet['C'+ str(l_new_rownum)].value  = l_tmp_value
            l_value = str(ref_sheet.cell(row=l_row_num, column=5).value)
            if l_value is not None:
                if 'DAY' in str(l_value).upper():
                    l_tmp_mnth = int(l_months_list[l_month]) - 1
                    if l_tmp_mnth < 1:
                        l_tmp_mnth = 12
                    l_Days =  calendar.monthrange(int(l_year),l_tmp_mnth)[1]
                    l_value = l_value.strip()
                    l_rate = int(l_value[5:])
                    l_amount = l_rate * l_Days
                else:
                    l_amount = l_value
            if l_amount != 'None':
                new_sheet['D'+ str(l_new_rownum)].value = int(l_amount)
            new_sheet['E'+ str(l_new_rownum)].value = 'Pending'

            new_sheet['G'+ str(l_new_rownum)].value =  ref_sheet.cell(row=l_row_num, column=7).value
    #print('done with sheet creation')
    return new_sheet
##~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~